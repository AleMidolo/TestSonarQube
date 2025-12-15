import openpyxl

class ExcelProcessor: 
    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        data = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            workbook.close()
            return data
        except:
            return None
    
    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return 1
        except:
            return 0
    
    def process_excel_data(self, N, save_file_name):
        """
        निर्दिष्ट कॉलम को Excel फ़ाइल में बड़े अक्षरों में बदलें
        :param N: int, उस कॉलम का अनुक्रमांक जिसे बदलना है
        :param save_file_name: str, स्रोत फ़ाइल का नाम
        :return:(int, str), पहला write_excel का लौटने वाला मान है, जबकि दूसरा संसाधित डेटा का सहेजा गया फ़ाइल नाम है
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        data = self.read_excel(save_file_name)
        if data is None:
            return 0, None
        
        for i in range(len(data)):
            row = list(data[i])
            if N < len(row):
                row[N] = row[N].upper() if isinstance(row[N], str) else row[N]
            data[i] = tuple(row)
        
        output_file_name = 'processed_' + save_file_name
        success = self.write_excel(data, output_file_name)
        return success, output_file_name