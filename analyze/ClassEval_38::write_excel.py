import openpyxl

class ExcelProcessor: 
    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        data = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            workbook.close()
            return data
        except:
            return None
    
    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0
        new_data = []
        for row in data:
            new_row = list(row[:])
            if not str(row[N]).isdigit():
                new_row[N] = str(row[N]).upper()
            new_data.append(new_row)
        new_file_name = save_file_name.split('.')[0] + '_process.xlsx'
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name

    def write_excel(self, data, file_name):
        """
        निर्दिष्ट Excel फ़ाइल में डेटा लिखें
        :param data: सूची, लिखने के लिए डेटा
        :param file_name: स्ट्रिंग, लिखने के लिए Excel फ़ाइल का नाम
        :return: 0 या 1, 1 सफल लेखन का प्रतिनिधित्व करता है, 0 असफल लेखन का प्रतिनिधित्व करता है
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('नाम', 'उम्र', 'देश'),
        >>>     ('जॉन', 25, 'यूएसए'),
        >>>     ('ऐलिस', 30, 'कनाडा'),
        >>>     ('बॉब', 35, 'ऑस्ट्रेलिया'),
        >>>     ('जूलिया', 28, 'जर्मनी')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return 1
        except:
            return 0