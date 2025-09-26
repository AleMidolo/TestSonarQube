import openpyxl


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        try:
            workbook = openpyxl.load_workbook(file_name)
            data = self._extract_data_from_sheet(workbook.active)
            workbook.close()
            return data
        except Exception:
            return None

    def _extract_data_from_sheet(self, sheet):
        return [row for row in sheet.iter_rows(values_only=True)]

    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            self._append_data_to_sheet(sheet, data)
            workbook.save(file_name)
            workbook.close()
            return 1
        except Exception:
            return 0

    def _append_data_to_sheet(self, sheet, data):
        for row in data:
            sheet.append(row)

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if self._is_invalid_data(data, N):
            return 0
        new_data = self._process_data(data, N)
        new_file_name = self._generate_new_file_name(save_file_name)
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name

    def _is_invalid_data(self, data, N):
        return data is None or N >= len(data[0])

    def _process_data(self, data, N):
        new_data = []
        for row in data:
            new_row = list(row[:])
            new_row.append(self._process_cell_value(row[N]))
            new_data.append(new_row)
        return new_data

    def _process_cell_value(self, value):
        return str(value).upper() if not str(value).isdigit() else value

    def _generate_new_file_name(self, original_file_name):
        return original_file_name.split('.')[0] + '_process.xlsx'